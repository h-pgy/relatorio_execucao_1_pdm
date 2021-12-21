import pandas as pd
from openpyxl import load_workbook
import os
from .utils import find_files_recursive, gen_col_range
from .config import (COL_INICIO,
                    COL_FIM,
                    HEADER_ROW,
                    DATA_ROW_RANGE,
                    DATA_SOURCE_DIR,
                    SAVE_CSV_DIR)

class XlDataReader:
    
    COL_INICIO = COL_INICIO
    COL_FIM = COL_FIM
    HEADER_ROW = HEADER_ROW
    DATA_ROW_RANGE = DATA_ROW_RANGE
    
    def __init__(self, lazy=True, file=None):

        self.lazy=lazy
        if not lazy and not file:
            raise ValueError(f'Se inicializar sem ser lazy, tem que passar o file')
        
        if not lazy:
            self.initialize(file=file)
            
    def list_data_sheets(self):
        
        data_sheets = [
            sheet
            for sheet in self.wb.sheetnames
            if sheet.lower().strip().endswith('indicador')
        ]
        
        return data_sheets
    
    
    def parse_sheet_name(self, name):
        
        new_name = name.lower().strip()
        
        splited = new_name.split(' ')
        
        secretaria = splited[0]
        meta = splited[1]
        meta = meta.replace('.', '_')
        
        return f'{secretaria}_{meta}'
    
    
    def parse_all_sheet_names(self, sheets):
        
        parsed = {self.parse_sheet_name(name) : name
                  for name in sheets}
        
        return parsed
    
    def get_sheet_names(self):
        
        data_sheets = self.list_data_sheets()
        
        return self.parse_all_sheet_names(data_sheets)
    
    def get_sheet(self, key):
        
        if key not in self.sheets:
            raise ValueError(f'Sheet {key} não encontrada no arquivo')
        
        return self.wb[self.sheets[key]]
    
    def get_cell_value(self, sheet, col, row):
        
        cell = f'{col}{row}'
        
        return sheet[cell].value
        
    def check_for_geo_data(self, sheet):
        
        #celula com o nome da primeira subprefeitura
        val = self.get_cell_value(
            sheet,
            self.COL_INICIO, 
            self.HEADER_ROW
        )
        
        if pd.isnull(val):
            return False
        elif val == '':
            return False
        val = val.lower().strip()
        
        if 'aricanduva' in val:
            return True
        return False
    
    
    def get_sub_name(self, sheet, col):
        
        
        val = self.get_cell_value(sheet,
                                   col, 
                                   self.HEADER_ROW)
        
        return val
    
    
    def parse_geo_data(self, sheet):
        
        
        col_range = self.gen_col_range(self.COL_INICIO, 
                                       self.COL_FIM)
        
        data = {}
        for col in col_range:
            sub_name = self.get_sub_name(sheet, col)
            data[sub_name] = []
            
            for row in range(*self.DATA_ROW_RANGE):
                
                val = self.get_cell_value(sheet, col, row)
                data[sub_name].append(val)
                
        return data
    
    def clean_geo_val(self, val):
        
        if pd.isnull(val):
            return float(0)
        if val == '':
            return float(0)
        
        val = str(val)
        val = val.replace('.', '')
        val = val.replace(',', '.')
        
        return float(val)
            
    
    def clean_geo_data(self, data):
        
        cleaned = {}
        
        for subs, vals in data.items():
            cleaned[subs] = []
            for val in vals:
                val = self.clean_geo_val(val)
                cleaned[subs].append(val)
                
        return cleaned
        
    
    def parse_all_geodata(self):
        
        all_data = {}
        for sheet_name in self.sheets.keys():
            sheet = self.get_sheet(sheet_name)
            if self.check_for_geo_data(sheet):
                data = self.parse_geo_data(sheet)
                clean_data = self.clean_geo_data(data)
                
                all_data[sheet_name] = clean_data
        
        return all_data

    def initialize(self, file):

        self.file = file
        self.gen_col_range = gen_col_range
        self.wb = load_workbook(file)
        self.sheets = self.get_sheet_names()

    def __call__(self, file=None):

        file = file or getattr(self,'file', None)
        if file is None:
            raise ValueError('Lazy load: precisa passar o parametro file')
        
        if self.lazy:
            self.initialize(file)

        return self.parse_all_geodata()

class DataParser:

    def __init__(self, source_folder = None, save_folder=None, overwrite=True):

        self.folder = source_folder or DATA_SOURCE_DIR
        self.save_folder = save_folder or SAVE_CSV_DIR
        self.overwrite = overwrite

        self.list_files = find_files_recursive
        self.read_xl = XlDataReader()

    def test_xl_file_name(self, file):
    
        fname = os.path.split(file)[-1]
        fname = fname.lower().strip()
        files_erradas = ('metas', 'iniciativas', 'controle')
        
        for teste in files_erradas:
            if fname.startswith(teste):
                return False
        
        return True

    def get_xl_files(self):
        
        xl_files = self.list_files(self.folder, extension='.xlsx')
        
        correct_files = []
        for file in xl_files:
            
            if self.test_xl_file_name(file):
                correct_files.append(file)
                
        
        return correct_files


    def parse_all_files(self, files=None):

        files = files or self.get_xl_files()
        all_data = []
        for file in files:
            print(f'Parsing {file}')
            data = self.read_xl(file)
            all_data.append(data)
            print('Parsed')
        
        return all_data

    def load_cached_file(self, output_file):

        file = os.path.join(self.save_folder, output_file)
        if os.path.exists(file):
            return pd.read_csv(file, sep=';', encoding='utf-8')

    def parse_sheet_name(self, sheet_name):

        splited = sheet_name.split('_')

        if len(splited) == 2:
            secretaria, meta = splited

            return secretaria, meta, ''
        else:
            secretaria, meta, obs = splited
            
            return secretaria, meta, obs

    def parse_to_df(self, sheet_name, data):

        df = pd.DataFrame(data)
        secretaria, meta, obs = self.parse_sheet_name(sheet_name)
        df['meta_num'] = meta
        df['secretaria'] = secretaria
        df['obs'] = obs
        df['file_name'] = sheet_name + '.jpg'


        return df

    def generate_dataframe(self, data):


        dfs = []
        for sheet_list in data:
            for sheet_name, data in sheet_list.items():

                df = self.parse_to_df(sheet_name, data)
                dfs.append(df)

        return pd.concat(dfs)

    def save_csv(self, df, output_file):

        file = os.path.join(self.save_folder, output_file)

        df.to_csv(file, sep=';', index=False, encoding='utf-8')

    def read_data(self, output_file='dados_regionalizacao_extraidos.csv'):

        if not self.overwrite:
            saved = self.load_cached_file(output_file)
            if saved is not None:
                print(f'Carregando arquivo já salvo: {output_file}')
                return saved
        
        data = self.parse_all_files()
        df = self.generate_dataframe(data)
        self.save_csv(df, output_file)

        return df

    def __call__(self):

        return self.read_data()
